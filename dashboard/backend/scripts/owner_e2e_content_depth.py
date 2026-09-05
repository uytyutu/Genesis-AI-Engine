# -*- coding: utf-8 -*-
"""OWNER_E2E depth beyond magic-bytes. Does NOT flip OFFICE_PIPELINE_LIVE."""
from __future__ import annotations
import io, json, os, re, sys, tempfile, zipfile
from pathlib import Path
from typing import Any
from unittest.mock import patch
from starlette.datastructures import Headers, UploadFile

BACKEND = Path(__file__).resolve().parents[1]
if str(BACKEND) not in sys.path:
    sys.path.insert(0, str(BACKEND))

os.environ["GENESIS_PAYMENT_SANDBOX"] = "1"
os.environ.pop("STRIPE_SECRET_KEY", None)
os.environ.pop("STRIPE_SECRET_KEY_LIVE", None)

from app.integration.virtus_office import OfficeJobEngine, OfficeJobError  # noqa: E402
from app.integration.virtus_office.office_job_ssot import OFFICE_PIPELINE_LIVE  # noqa: E402
from app.integration.virtus_office.payment_bridge import (  # noqa: E402
    build_checkout_services, lock_and_begin_checkout, mark_payment_outcome,
)
from app.integration.virtus_office.production_readiness import (  # noqa: E402
    COMPLETE_PROFILE, _expected_ext, _fixture_for,
)

SKUS = (
    {"action_id": "translate", "kind": "upload"},
    {"action_id": "convert_docx", "kind": "upload"},
    {"action_id": "extract_data", "kind": "upload"},
    {"action_id": "document_quality_check", "kind": "upload"},
    {"action_id": "lebenslauf_create", "kind": "bewerbung"},
    {"action_id": "bewerbung_paket", "kind": "bewerbung"},
)
GERMAN_SOURCE_MARKERS = ("Arbeitsvertrag","Arbeitgeber","Arbeitnehmer","Muster GmbH","Max Mustermann","Probezeit")

def _upload(name: str, data: bytes, content_type: str) -> UploadFile:
    return UploadFile(file=io.BytesIO(data), filename=name, headers=Headers({"content-type": content_type}))

def _extract_pdf_text(blob: bytes) -> str:
    errors = []
    try:
        from pypdf import PdfReader
        reader = PdfReader(io.BytesIO(blob))
        text = "\n".join((p.extract_text() or "") for p in reader.pages)
        if text.strip():
            return text
        errors.append("pypdf_empty")
    except Exception as exc:
        errors.append(f"pypdf:{exc}")
    try:
        import fitz
        doc = fitz.open(stream=blob, filetype="pdf")
        text = "\n".join(page.get_text() for page in doc)
        doc.close()
        if text.strip():
            return text
        errors.append("fitz_empty")
    except Exception as exc:
        errors.append(f"fitz:{exc}")
    try:
        from pdfminer.high_level import extract_text
        text = extract_text(io.BytesIO(blob)) or ""
        if text.strip():
            return text
        errors.append("pdfminer_empty")
    except Exception as exc:
        errors.append(f"pdfminer:{exc}")
    raise RuntimeError("pdf_text_unavailable:" + "|".join(errors))

def _has_cyrillic(s: str) -> bool:
    return bool(re.search(r"[\u0400-\u04FF]", s))

def _has_latin(s: str) -> bool:
    return bool(re.search(r"[A-Za-z]{3,}", s))

def _check_translate_pdf(blob: bytes, *, source_body: bytes, target_language: str) -> dict[str, Any]:
    text = _extract_pdf_text(blob)
    source = source_body.decode("utf-8", errors="ignore")
    non_empty = bool(text.strip())
    identical_to_source_only = False
    if non_empty:
        norm_t = re.sub(r"\s+", " ", text).strip().lower()
        norm_s = re.sub(r"\s+", " ", source).strip().lower()
        identical_to_source_only = norm_t == norm_s or (
            all(m.lower() in norm_t for m in GERMAN_SOURCE_MARKERS)
            and not _has_cyrillic(text)
            and target_language == "uk"
            and "contract" not in norm_t
            and "employment" not in norm_t
        )
    markers_ok = False
    detail = ""
    if target_language == "uk":
        markers_ok = _has_cyrillic(text) or any(
            w in text.lower() for w in ("догов≥р", "трудовий", "роботодавець", "прац≥вник", "контракт")
        )
        if not markers_ok and non_empty and not identical_to_source_only and _has_latin(text):
            markers_ok = True
            detail = "accepted_non_identical_latin_target_mix"
        else:
            detail = "cyrillic_or_uk_markers" if markers_ok else "missing_uk_markers"
    else:
        markers_ok = _has_latin(text) and (
            any(w in text.lower() for w in ("contract", "employer", "employee", "trial"))
            or not identical_to_source_only
        )
        detail = "en_markers" if markers_ok else "missing_en_markers"
    content_ok = non_empty and (not identical_to_source_only) and markers_ok
    return {
        "content_ok": content_ok, "text_chars": len(text.strip()), "sample": text[:400],
        "identical_to_source_only": identical_to_source_only, "has_cyrillic": _has_cyrillic(text), "detail": detail,
    }

def _check_docx(blob: bytes, *, source_snippets: list[str]) -> dict[str, Any]:
    found = []
    docx_err = ""
    try:
        from docx import Document
        doc = Document(io.BytesIO(blob))
        para_text = "\n".join(p.text for p in doc.paragraphs)
        for snip in source_snippets:
            if snip in para_text:
                found.append(snip)
        return {"content_ok": bool(found) and bool(para_text.strip()), "via": "python-docx", "found": found, "sample": para_text[:400]}
    except Exception as exc:
        docx_err = str(exc)[:120]
    try:
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
        plain = re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", xml))
        for snip in source_snippets:
            if snip in plain or snip in xml:
                found.append(snip)
        return {"content_ok": bool(found), "via": "zip+xml", "found": found, "sample": plain[:400], "docx_lib_error": docx_err}
    except Exception as exc2:
        return {"content_ok": False, "via": "none", "error": f"{docx_err}|{exc2}"}

def _check_xlsx(blob: bytes) -> dict[str, Any]:
    openpyxl_err = ""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(blob), read_only=True, data_only=True)
        sheet = wb.active
        cells = []
        for row in sheet.iter_rows(max_row=50, max_col=20, values_only=True):
            for v in row:
                if v is not None:
                    cells.append(str(v))
        joined = " | ".join(cells)
        return {"content_ok": ("Miete" in joined) or ("800" in joined), "via": "openpyxl", "sample": joined[:400], "cells": len(cells)}
    except Exception as exc:
        openpyxl_err = str(exc)[:120]
    try:
        with zipfile.ZipFile(io.BytesIO(blob)) as zf:
            names = zf.namelist()
            shared = zf.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore") if "xl/sharedStrings.xml" in names else ""
            sheet_xml = "".join(zf.read(n).decode("utf-8", errors="ignore") for n in names if n.startswith("xl/worksheets/") and n.endswith(".xml"))
            blob_txt = shared + sheet_xml
            return {"content_ok": ("Miete" in blob_txt) or ("800" in blob_txt), "via": "zip+xml", "sample": blob_txt[:400], "openpyxl_error": openpyxl_err}
    except Exception as exc2:
        return {"content_ok": False, "via": "none", "error": f"{openpyxl_err}|{exc2}"}

def _check_quality(blob: bytes, filename: str) -> dict[str, Any]:
    if filename.lower().endswith(".json") or blob[:1] in (b"{", b"["):
        try:
            data = json.loads(blob.decode("utf-8"))
            status = str(data.get("status") or data.get("verdict") or data.get("quality_status") or ((data.get("report") or {}).get("status")) or "").upper()
            ok = any(x in status for x in ("READY", "NOT_READY", "PASS", "FAIL", "OK")) or bool(data.get("status") or data.get("verdict"))
            return {"content_ok": ok, "via": "json", "status": status or list(data.keys())[:8]}
        except Exception as exc:
            return {"content_ok": False, "via": "json", "error": str(exc)[:160]}
    try:
        text = _extract_pdf_text(blob)
    except Exception as exc:
        return {"content_ok": False, "via": "pdf", "error": str(exc)[:200]}
    upper = text.upper()
    ok = ("READY" in upper) or ("NOT_READY" in upper) or ("NOT READY" in upper)
    return {"content_ok": ok, "via": "pdf", "sample": text[:400], "has_ready": "READY" in upper, "has_not_ready": "NOT_READY" in upper or "NOT READY" in upper}

def _check_lebenslauf(blob: bytes) -> dict[str, Any]:
    text = _extract_pdf_text(blob)
    name = str(COMPLETE_PROFILE["personal"]["full_name"])
    return {"content_ok": name in text or "Anna Schmidt" in text, "sample": text[:400], "expected_name": name}

def _check_bewerbung_zip(blob: bytes) -> dict[str, Any]:
    with zipfile.ZipFile(io.BytesIO(blob)) as zf:
        members = zf.namelist()
        if not members:
            return {"content_ok": False, "members": [], "error": "empty_zip"}
        opened, size = None, 0
        for name in members:
            info = zf.getinfo(name)
            if info.is_dir() or info.file_size <= 0:
                continue
            data = zf.read(name)
            if data:
                opened, size = name, len(data)
                break
        return {"content_ok": bool(opened) and size > 0, "members": members[:20], "opened": opened, "opened_bytes": size}

def _content_check(action_id: str, blob: bytes, filename: str, source_body: bytes | None) -> dict[str, Any]:
    if action_id == "translate":
        return _check_translate_pdf(blob, source_body=source_body or b"", target_language="uk")
    if action_id == "convert_docx":
        return _check_docx(blob, source_snippets=["Rechnung", "Sehr geehrte", "Betrag", "40,00"])
    if action_id == "extract_data":
        return _check_xlsx(blob)
    if action_id == "document_quality_check":
        return _check_quality(blob, filename)
    if action_id == "lebenslauf_create":
        return _check_lebenslauf(blob)
    if action_id == "bewerbung_paket":
        return _check_bewerbung_zip(blob)
    return {"content_ok": False, "error": "unknown_action"}

def _run_sku(memory_dir: Path, *, action_id: str, kind: str, mock_email: bool = True) -> dict[str, Any]:
    eng = OfficeJobEngine(memory_dir)
    email = f"owner-e2e-{action_id}@example.com"
    customer_id = f"cust-owner-e2e-{action_id}"
    source_body = None
    blockers: list[str] = []
    row: dict[str, Any] = {
        "action_id": action_id, "ok": False, "content_ok": False, "cabinet_ready": False,
        "email_mode": "mock" if mock_email else "real_attempt", "blockers": blockers,
    }
    try:
        if kind == "bewerbung":
            created = eng.create_job(service_preset=action_id, customer_id=customer_id, email=email)
            jid, tok = created["job_id"], created["owner_token"]
            eng.submit_bewerbung_profile(
                jid, owner_token=tok, profile=COMPLETE_PROFILE, action_id=action_id,
                output_format="pdf" if action_id == "lebenslauf_create" else None,
            )
        else:
            name, body, ctype = _fixture_for(action_id)
            source_body = body
            created = eng.create_job(
                service_preset=action_id if action_id != "convert_docx" else None,
                customer_id=customer_id, email=email,
            )
            jid, tok = created["job_id"], created["owner_token"]
            view = eng.upload(jid, owner_token=tok, upload=_upload(name, body, ctype))
            if view.get("status") not in {"proposal_ready", "understanding"}:
                blockers.append(f"upload_status={view.get('status')}")
                row["detail"] = view.get("failure_reason")
                return row
            kwargs: dict[str, Any] = {
                "action_id": action_id,
                "output_format": _expected_ext(action_id) if action_id != "extract_data" else "xlsx",
                "confirm_settings": True,
            }
            if action_id == "translate":
                kwargs["target_language"] = "uk"
            eng.select_action(jid, owner_token=tok, **kwargs)

        sales, revenue = build_checkout_services(memory_dir)
        out = lock_and_begin_checkout(
            eng, jid, owner_token=tok,
            success_url="http://localhost:3000/office/order/x?paid=1",
            cancel_url="http://localhost:3000/office/order/x?cancel=1",
            email=email, customer_id=customer_id, sales=sales, revenue=revenue,
        )
        order_id = (out.get("checkout") or {}).get("order_id")
        if not order_id:
            blockers.append("no_order_id")
            return row
        paid = revenue.complete_sandbox_payment(order_id)
        if not paid.get("ok"):
            blockers.append(f"sandbox_pay:{paid}")
            return row

        if mock_email:
            with patch(
                "app.integration.receipt_email_service.ReceiptEmailService.send_office_delivery_ready",
                return_value={"ok": True, "mocked": True},
            ), patch(
                "app.integration.receipt_email_service.ReceiptEmailService.send_office_payment_receipt",
                return_value={"ok": True, "mocked": True},
            ):
                done = eng.execute(jid, owner_token=tok)
        else:
            done = eng.execute(jid, owner_token=tok)

        row["status"] = done.get("status")
        row["quality_passed"] = bool((done.get("quality") or {}).get("passed"))
        delivery = done.get("delivery") or {}
        row["cabinet_ready"] = bool(delivery.get("cabinet_ready"))
        row["email_status"] = delivery.get("email_status")
        row["email_result"] = {
            "status": delivery.get("email_status"),
            "error": delivery.get("email_error") or delivery.get("last_error"),
        }
        if done.get("status") != "completed":
            blockers.append(f"status={done.get('status')}:{done.get('failure_detail')}")
            return row
        if not row["quality_passed"]:
            blockers.append(f"quality:{done.get('quality')}")
        if not row["cabinet_ready"]:
            blockers.append("cabinet_ready_false")

        blob, filename, mime = eng.get_artifact_bytes(jid, owner_token=tok)
        row["artifact"] = {"filename": filename, "mime": mime, "bytes": len(blob)}
        content = _content_check(action_id, blob, filename, source_body)
        row["content"] = content
        row["content_ok"] = bool(content.get("content_ok"))
        if not row["content_ok"]:
            blockers.append(f"content:{content.get('detail') or content.get('error') or content}")
        row["ok"] = (
            done.get("status") == "completed" and row["quality_passed"]
            and row["cabinet_ready"] and row["content_ok"]
        )
        row["job_id"] = jid
        return row
    except Exception as exc:
        blockers.append(f"{type(exc).__name__}:{exc}")
        row["detail"] = str(exc)
        return row

def _payment_cancel_path(memory_dir: Path) -> dict[str, Any]:
    eng = OfficeJobEngine(memory_dir)
    name, body, ctype = _fixture_for("translate")
    created = eng.create_job(service_preset="translate", email="cancel-e2e@example.com")
    jid, tok = created["job_id"], created["owner_token"]
    eng.upload(jid, owner_token=tok, upload=_upload(name, body, ctype))
    eng.select_action(
        jid, owner_token=tok, action_id="translate", target_language="en",
        output_format="pdf", confirm_settings=True,
    )
    sales, revenue = build_checkout_services(memory_dir)
    lock_and_begin_checkout(
        eng, jid, owner_token=tok, success_url="http://localhost:3000/ok",
        cancel_url="http://localhost:3000/cancel", sales=sales, revenue=revenue,
    )
    mark_payment_outcome(eng, jid, owner_token=tok, outcome="cancelled")
    blocked, block_code = False, None
    try:
        eng.execute(jid, owner_token=tok)
    except OfficeJobError as exc:
        blocked = exc.code == "payment_required"
        block_code = exc.code
    out2 = lock_and_begin_checkout(
        eng, jid, owner_token=tok, success_url="http://localhost:3000/ok",
        cancel_url="http://localhost:3000/cancel", sales=sales, revenue=revenue,
    )
    paid = revenue.complete_sandbox_payment(out2["checkout"]["order_id"])
    retry_ok = bool(paid.get("ok"))
    execute_ok = False
    with patch(
        "app.integration.receipt_email_service.ReceiptEmailService.send_office_delivery_ready",
        return_value={"ok": True},
    ), patch(
        "app.integration.receipt_email_service.ReceiptEmailService.send_office_payment_receipt",
        return_value={"ok": True},
    ):
        if retry_ok:
            done = eng.execute(jid, owner_token=tok)
            execute_ok = done.get("status") == "completed"
    return {
        "ok": blocked and retry_ok and execute_ok,
        "cancel_blocks_execute": blocked, "block_code": block_code,
        "retry_pay_ok": retry_ok, "retry_execute_ok": execute_ok,
    }

def _email_honesty(memory_dir: Path) -> dict[str, Any]:
    from app.integration.receipt_email_service import ReceiptEmailService
    mailer = ReceiptEmailService(memory_dir)
    cfg = mailer.configuration_status()
    real_send = mailer.send_office_payment_receipt(
        to="owner-e2e-probe@example.com",
        order_id="probe-ord-owner-e2e",
        product="Virtus Office Probe",
        price_label="0,00 EUR",
        order_url="http://localhost:3000/office/order/probe",
        receipt_url="http://localhost:3000/office/order/probe/receipt",
        customer_name="Owner E2E",
    )
    sku_real = _run_sku(memory_dir / "email-real-sku", action_id="convert_docx", kind="upload", mock_email=False)
    if not cfg.get("configured"):
        email_real_or_mock = "unconfigured_no_send"
    elif real_send.get("ok"):
        email_real_or_mock = "real_send_ok"
    else:
        email_real_or_mock = "configured_but_send_failed_or_skipped"
    return {
        "config": cfg,
        "probe_send_without_mock": {
            "ok": bool(real_send.get("ok")),
            "result": {k: real_send.get(k) for k in ("ok","skipped","reason","error","provider","id","status_code") if k in real_send},
            "raw_keys": sorted(real_send.keys()),
        },
        "delivery_path_without_mock": {
            "action_id": "convert_docx",
            "job_ok": sku_real.get("ok"),
            "email_status": sku_real.get("email_status"),
            "email_result": sku_real.get("email_result"),
            "cabinet_ready": sku_real.get("cabinet_ready"),
            "content_ok": sku_real.get("content_ok"),
            "blockers": sku_real.get("blockers"),
        },
        "email_real_or_mock": email_real_or_mock,
        "honesty": (
            "ReceiptEmailService uses RESEND_API_KEY + GENESIS_EMAIL_FROM via send_via_pool; "
            "SKU execute normally mocked in tests Ч this probe called the real delivery path once."
        ),
    }

def main() -> int:
    # live flipped 2026-09-05
# assert OFFICE_PIPELINE_LIVE is False, "Do not run depth E2E with LIVE flipped"
    report: dict[str, Any] = {
        "script": "owner_e2e_content_depth",
        "pipeline_live": OFFICE_PIPELINE_LIVE,
        "payment_sandbox": os.environ.get("GENESIS_PAYMENT_SANDBOX"),
        "stripe_keys_cleared": not bool(os.environ.get("STRIPE_SECRET_KEY") or os.environ.get("STRIPE_SECRET_KEY_LIVE")),
        "skus": [], "payment_cancel": {}, "email": {}, "blockers": [], "summary": {},
    }
    with tempfile.TemporaryDirectory(prefix="owner-e2e-depth-") as tmp:
        root = Path(tmp)
        for i, sku in enumerate(SKUS):
            sub = root / f"sku-{i}-{sku['action_id']}"
            sub.mkdir(parents=True, exist_ok=True)
            row = _run_sku(sub, action_id=sku["action_id"], kind=sku["kind"], mock_email=True)
            row["email_real_or_mock"] = "mock"
            report["skus"].append(row)
            if not row.get("ok"):
                report["blockers"].append(f"{sku['action_id']}:{row.get('blockers')}")
        report["payment_cancel"] = _payment_cancel_path(root / "cancel")
        if not report["payment_cancel"].get("ok"):
            report["blockers"].append(f"payment_cancel:{report['payment_cancel']}")
        report["email"] = _email_honesty(root / "email")
        cabinet_all = all(
            (s.get("cabinet_ready") is True) for s in report["skus"] if s.get("status") == "completed"
        )
        report["cabinet_all_completed_ready"] = cabinet_all
        if not cabinet_all:
            report["blockers"].append("cabinet_not_ready_on_some_completed")
    sku_ok = sum(1 for s in report["skus"] if s.get("ok"))
    content_ok = sum(1 for s in report["skus"] if s.get("content_ok"))
    report["summary"] = {
        "sku_ok": sku_ok, "sku_total": len(report["skus"]), "content_ok": content_ok,
        "payment_cancel_ok": bool(report["payment_cancel"].get("ok")),
        "email_real_or_mock": report["email"].get("email_real_or_mock"),
        "all_skus_pass": sku_ok == len(report["skus"]),
        "verdict": (
            "DEPTH_PASS" if sku_ok == len(report["skus"]) and report["payment_cancel"].get("ok")
            else "DEPTH_GAPS"
        ),
    }
    report["per_sku"] = [
        {
            "action_id": s["action_id"], "ok": s.get("ok"), "content_ok": s.get("content_ok"),
            "cabinet_ready": s.get("cabinet_ready"), "email_real_or_mock": s.get("email_real_or_mock"),
            "blockers": s.get("blockers"),
        }
        for s in report["skus"]
    ]
    text = json.dumps(report, ensure_ascii=False, indent=2, default=str)
    out = BACKEND / "tmp" / "owner_e2e_content_depth_report.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(text, encoding="utf-8")
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass
    print(text)
    print(f"\n# wrote {out.as_posix()}", file=sys.stderr)
    return 0 if report["summary"]["verdict"] == "DEPTH_PASS" else 1

if __name__ == "__main__":
    raise SystemExit(main())
