"""File parsing strategies for Virtus Office — text layer + Stage 4 OCR.

Reuses knowledge_intake_pdf for PDF text. Scans/images go through ocr_engine.
No second storage.
"""

from __future__ import annotations

import csv
import io
import zipfile
from typing import Any
from xml.etree import ElementTree as ET

from app.integration.knowledge_intake_pdf import extract_pdf_text_bytes
from app.integration.virtus_office.document_classify import count_tables_heuristic
from app.integration.virtus_office.ocr_engine import (
    layout_from_plain_text,
    ocr_image_bytes,
    ocr_image_pages,
    ocr_pdf_bytes,
)


def _pdf_image_count(data: bytes) -> int:
    try:
        from pypdf import PdfReader

        reader = PdfReader(io.BytesIO(data))
        imgs = 0
        for page in reader.pages:
            res = page.get("/Resources")
            if not res:
                continue
            res = res.get_object() if hasattr(res, "get_object") else res
            xo = res.get("/XObject") if res else None
            if not xo:
                continue
            xo = xo.get_object() if hasattr(xo, "get_object") else xo
            for key in xo:
                obj = xo[key].get_object()
                if obj.get("/Subtype") == "/Image":
                    imgs += 1
        return imgs
    except Exception:
        return 0


def _docx_text(data: bytes) -> str:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            raw = zf.read("word/document.xml")
        root = ET.fromstring(raw)
        texts = [
            (node.text or "")
            for node in root.iter()
            if node.tag.endswith("}t") and node.text
        ]
        return " ".join(texts).strip()
    except Exception:
        return ""


def _docx_image_count(data: bytes) -> int:
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            return sum(1 for n in zf.namelist() if n.startswith("word/media/"))
    except Exception:
        return 0


def _xlsx_preview(data: bytes) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook  # type: ignore

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        sheets = list(wb.sheetnames)
        rows_sample: list[list[str]] = []
        if sheets:
            ws = wb[sheets[0]]
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= 8:
                    break
                rows_sample.append([str(c) if c is not None else "" for c in row[:12]])
        wb.close()
        flat = "\n".join(" ".join(r) for r in rows_sample)
        return {
            "text": flat[:8000],
            "sheet_count": len(sheets),
            "sheets": sheets[:20],
            "tables": max(1, len(sheets)),
            "parse_ok": True,
        }
    except Exception:
        pass
    try:
        with zipfile.ZipFile(io.BytesIO(data)) as zf:
            names = zf.namelist()
            sheets = [n for n in names if n.startswith("xl/worksheets/")]
            text = ""
            if "xl/sharedStrings.xml" in names:
                raw = zf.read("xl/sharedStrings.xml").decode("utf-8", errors="ignore")
                text = " ".join(
                    m.group(1)
                    for m in __import__("re").finditer(r"<t[^>]*>([^<]+)</t>", raw)
                )[:8000]
            return {
                "text": text,
                "sheet_count": len(sheets) or 1,
                "sheets": [],
                "tables": max(1, len(sheets) or 1),
                "parse_ok": bool(text) or bool(sheets),
            }
    except Exception:
        return {
            "text": "",
            "sheet_count": 0,
            "sheets": [],
            "tables": 0,
            "parse_ok": False,
        }


def _image_meta(data: bytes) -> dict[str, Any]:
    try:
        from PIL import Image  # type: ignore

        im = Image.open(io.BytesIO(data))
        w, h = im.size
        return {"width": w, "height": h, "format": (im.format or "").lower(), "parse_ok": True}
    except Exception:
        return {"width": None, "height": None, "format": None, "parse_ok": False}


def _apply_ocr_success(base: dict[str, Any], ocr: dict[str, Any], *, strategy: str) -> dict[str, Any]:
    text = str(ocr.get("text") or "")
    pages = list(ocr.get("pages") or [])
    tables = list(ocr.get("tables") or [])
    base.update(
        {
            "text": text,
            "text_detected": bool(text.strip()),
            "page_count": len(pages) or base.get("page_count") or 1,
            "pages_included": len(pages) or base.get("pages_included"),
            "tables": max(int(base.get("tables") or 0), len(tables), count_tables_heuristic(text)),
            "text_blocks": len([p for p in text.split("\n\n") if p.strip()]) if text else 0,
            "ocr_status": "done",
            "parse_ok": True,
            "parse_strategy": strategy,
            "ocr": {
                "provider": ocr.get("provider"),
                "confidence": ocr.get("confidence"),
                "language": ocr.get("language"),
                "page_count": len(pages),
                "has_layout": bool(pages),
                "table_count": len(tables),
                "financial_qa": ocr.get("financial_qa"),
                "needs_review": bool(ocr.get("needs_review")),
                "review_warning_de": ocr.get("review_warning_de"),
            },
            "layout": {
                "pages": pages,
                "tables": tables,
                "lines": [ln for pg in pages for ln in (pg.get("lines") or [])],
                "blocks": [b for pg in pages for b in (pg.get("blocks") or [])],
            },
        }
    )
    base["notes"].append(f"ocr_provider:{ocr.get('provider')}")
    fq = ocr.get("financial_qa") or {}
    if fq.get("needs_review"):
        base["notes"].append("ocr_financial_needs_review")
        base["financial_qa"] = fq
    elif fq.get("is_financial"):
        base["financial_qa"] = fq
    return base


def _apply_ocr_failure(base: dict[str, Any], ocr: dict[str, Any]) -> dict[str, Any]:
    err = str(ocr.get("error") or "ocr_failed")
    base["ocr_status"] = "failed"
    base["parse_ok"] = bool(base.get("parse_ok"))
    base["ocr"] = {
        "provider": ocr.get("provider") or "none",
        "confidence": 0.0,
        "language": None,
        "error": err,
        "detail": ocr.get("detail"),
    }
    base["layout"] = {"pages": [], "tables": [], "lines": [], "blocks": []}
    base["notes"].append(f"ocr_failed:{err}")
    return base


def parse_office_file(
    *,
    data: bytes,
    filename: str,
    file_kind: str,
    content_type: str = "",
    extra_pages: list[tuple[bytes, str]] | None = None,
) -> dict[str, Any]:
    """Return structure + text (+ OCR layout when needed).

    extra_pages: additional image pages (PNG/JPEG bytes, content_type) for multi-shot.
    """
    kind = file_kind or "unknown"
    base: dict[str, Any] = {
        "file_kind": kind,
        "filename": filename,
        "content_type": content_type,
        "text": "",
        "text_detected": False,
        "page_count": None,
        "pages_included": None,
        "tables": 0,
        "images": 0,
        "text_blocks": 0,
        "sheet_count": None,
        "ocr_status": "not_needed",
        "parse_strategy": kind,
        "parse_ok": False,
        "notes": [],
        "ocr": None,
        "layout": None,
    }

    extras = list(extra_pages or [])

    if kind == "pdf":
        try:
            # Full Office jobs: all pages, no Knowledge Intake 14k char trim.
            text, total, included = extract_pdf_text_bytes(
                data, max_pages=80, max_chars=None
            )
        except Exception as exc:
            base["notes"].append(f"pdf_parse_error:{exc}")
            # Try OCR raster path
            ocr = ocr_pdf_bytes(data, max_pages=40)
            if ocr.get("ok"):
                return _apply_ocr_success(base, ocr, strategy="pdf_ocr_after_parse_error")
            return _apply_ocr_failure(base, ocr)

        base["text"] = text
        base["text_detected"] = bool(text.strip())
        base["page_count"] = total
        base["pages_included"] = included
        base["images"] = _pdf_image_count(data)
        base["tables"] = count_tables_heuristic(text)
        base["text_blocks"] = len([p for p in text.split("\n\n") if p.strip()]) if text else 0
        base["parse_ok"] = True
        base["parse_strategy"] = "knowledge_intake_pdf"
        if text.strip():
            layout = layout_from_plain_text(text)
            base["layout"] = {
                "pages": layout.get("pages") or [],
                "tables": [],
                "lines": [ln for pg in (layout.get("pages") or []) for ln in (pg.get("lines") or [])],
                "blocks": [b for pg in (layout.get("pages") or []) for b in (pg.get("blocks") or [])],
            }
            return base

        # Scanned / image PDF — Stage 4 OCR
        base["notes"].append("scanned_or_image_pdf_no_text_layer")
        ocr = ocr_pdf_bytes(data, max_pages=40)
        if ocr.get("ok"):
            return _apply_ocr_success(base, ocr, strategy="pdf_scan_ocr")
        return _apply_ocr_failure(base, ocr)

    if kind == "docx":
        text = _docx_text(data)
        imgs = _docx_image_count(data)
        base["text"] = text
        base["text_detected"] = bool(text.strip())
        base["page_count"] = 1 if text else None
        base["images"] = imgs
        base["tables"] = count_tables_heuristic(text)
        base["text_blocks"] = len([p for p in text.split("\n") if p.strip()]) if text else 0
        base["parse_ok"] = True
        base["parse_strategy"] = "docx_xml_reuse"
        if text.strip():
            layout = layout_from_plain_text(text)
            base["layout"] = {
                "pages": layout.get("pages") or [],
                "tables": [],
                "lines": [ln for pg in (layout.get("pages") or []) for ln in (pg.get("lines") or [])],
                "blocks": [b for pg in (layout.get("pages") or []) for b in (pg.get("blocks") or [])],
            }
        return base

    if kind == "xlsx":
        prev = _xlsx_preview(data)
        base.update(
            {
                "text": prev.get("text") or "",
                "text_detected": bool((prev.get("text") or "").strip()),
                "tables": int(prev.get("tables") or 0),
                "sheet_count": prev.get("sheet_count"),
                "parse_ok": bool(prev.get("parse_ok")),
                "parse_strategy": "xlsx_openpyxl_or_zip",
                "notes": [f"sheets:{','.join(prev.get('sheets') or [])}"[:120]],
            }
        )
        return base

    if kind == "csv":
        try:
            raw = data.decode("utf-8", errors="ignore")
            sample = raw[:12000]
            reader = csv.reader(io.StringIO(sample))
            rows = list(reader)[:40]
            flat = "\n".join(",".join(r) for r in rows)
            base.update(
                {
                    "text": flat,
                    "text_detected": bool(flat.strip()),
                    "tables": 1 if rows else 0,
                    "sheet_count": 1,
                    "parse_ok": True,
                    "parse_strategy": "csv",
                    "text_blocks": len(rows),
                }
            )
        except Exception as exc:
            base["notes"].append(f"csv_error:{exc}")
        return base

    if kind == "txt":
        text = data.decode("utf-8", errors="ignore")
        base.update(
            {
                "text": text[:14000],
                "text_detected": bool(text.strip()),
                "page_count": 1,
                "text_blocks": len([p for p in text.split("\n\n") if p.strip()]),
                "parse_ok": True,
                "parse_strategy": "txt",
            }
        )
        if text.strip():
            layout = layout_from_plain_text(text[:14000])
            base["layout"] = {
                "pages": layout.get("pages") or [],
                "tables": [],
                "lines": [ln for pg in (layout.get("pages") or []) for ln in (pg.get("lines") or [])],
                "blocks": [b for pg in (layout.get("pages") or []) for b in (pg.get("blocks") or [])],
            }
        return base

    if kind == "image":
        meta = _image_meta(data)
        base.update(
            {
                "images": 1 + len(extras),
                "page_count": 1 + len(extras),
                "parse_ok": bool(meta.get("parse_ok")),
                "image_meta": meta,
                "notes": [f"size:{meta.get('width')}x{meta.get('height')}"],
            }
        )
        page_blobs: list[tuple[bytes, str]] = [
            (data, content_type or "image/png"),
            *[(b, ct or "image/png") for b, ct in extras],
        ]
        if len(page_blobs) == 1:
            ocr = ocr_image_bytes(page_blobs[0][0], content_type=page_blobs[0][1])
        else:
            ocr = ocr_image_pages(page_blobs)
        if ocr.get("ok"):
            return _apply_ocr_success(base, ocr, strategy="image_ocr")
        return _apply_ocr_failure(base, ocr)

    base["notes"].append("unsupported_parse_kind")
    return base
